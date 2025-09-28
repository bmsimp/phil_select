#!/usr/bin/env python3

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json

def analyze_excel_file(filepath):
    """Analyze Excel file structure and extract information about worksheets and data"""
    
    print(f"Analyzing Excel file: {filepath}")
    print("=" * 60)
    
    # Load workbook to get worksheet names and examine structure
    workbook = load_workbook(filepath, data_only=False)
    
    print(f"Number of worksheets: {len(workbook.worksheets)}")
    print("Worksheet names:")
    
    worksheet_info = {}
    
    for i, sheet in enumerate(workbook.worksheets):
        print(f"  {i+1}. {sheet.title}")
        
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        worksheet_info[sheet.title] = {
            'max_row': max_row,
            'max_col': max_col,
            'has_data': max_row > 1 or max_col > 1
        }
        
        print(f"     Dimensions: {max_row} rows x {max_col} columns")
        
        # Sample first few rows to understand structure
        if max_row > 0 and max_col > 0:
            print(f"     First few rows:")
            for row_num in range(1, min(6, max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, max_col + 1)):  # First 10 columns
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    if value is not None:
                        row_data.append(str(value)[:30])  # Truncate long values
                    else:
                        row_data.append("")
                if any(row_data):  # Only print if row has data
                    print(f"       Row {row_num}: {row_data}")
        
        print()
    
    # Check for macros/VBA code
    print("Checking for VBA macros...")
    if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
        print("  ✓ VBA macros detected in the file")
        print("  Note: VBA code cannot be directly extracted with openpyxl")
    else:
        print("  No VBA macros detected (or they're not accessible)")
    
    print("\nNamed ranges:")
    for name, range_obj in workbook.defined_names.items():
        print(f"  {name}: {range_obj.attr_text}")
    
    return worksheet_info

def extract_data_tables(filepath):
    """Extract data from each worksheet as pandas DataFrames"""
    print("\nExtracting data tables...")
    print("=" * 40)
    
    # Get all sheet names first
    xl_file = pd.ExcelFile(filepath)
    
    data_tables = {}
    
    for sheet_name in xl_file.sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        try:
            # Try to read the sheet
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            
            print(f"  Shape: {df.shape}")
            print(f"  Columns: {list(df.columns)}")
            
            # Show sample data
            if not df.empty:
                print(f"  Sample data:")
                print(df.head(3).to_string())
            
            data_tables[sheet_name] = {
                'dataframe': df,
                'shape': df.shape,
                'columns': list(df.columns),
                'dtypes': df.dtypes.to_dict()
            }
            
        except Exception as e:
            print(f"  Error reading sheet: {e}")
    
    return data_tables

if __name__ == "__main__":
    filepath = "/home/brian/git/philmont_selection/treks.xlsm"
    
    # Analyze structure
    worksheet_info = analyze_excel_file(filepath)
    
    # Extract data
    data_tables = extract_data_tables(filepath)
    
    # Save analysis results
    analysis_results = {
        'worksheet_info': worksheet_info,
        'table_summaries': {}
    }
    
    for sheet_name, table_info in data_tables.items():
        analysis_results['table_summaries'][sheet_name] = {
            'shape': table_info['shape'],
            'columns': table_info['columns'],
            'dtypes': {k: str(v) for k, v in table_info['dtypes'].items()}
        }
    
    with open('excel_analysis.json', 'w') as f:
        json.dump(analysis_results, f, indent=2)
    
    print(f"\nAnalysis complete! Results saved to excel_analysis.json")